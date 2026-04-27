"""
Generador de migración 00007_seed_deudas_iniciales.sql
Aplica los parámetros del checkpoint de calidad:
  • Normaliza floats → int en códigos de puesto
  • Usa 'Hoja1' como fuente canónica para amperaje (0.75 S/kWh socios, 0.77 inquilinos)
  • Detecta dias_uso=40 u horas_uso=14 via ratio KWH_excel/KWH_calc
  • Agua sin código de puesto: cruce por apellidos en SQL (socios + inquilinos)
"""

import openpyxl
import re
import json
import sys
import io
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

BASE = Path(r"c:/Github/SistemaCooperativa/datos_iniciales")

# ── helpers ───────────────────────────────────────────────────────────────────

def norm_puesto(s):
    """Normaliza código de puesto. Convierte 13.0 → '13', maneja guiones."""
    if s is None:
        return ''
    # float como 13.0 → '13'
    try:
        f = float(str(s))
        if f == int(f) and f > 0:
            return str(int(f))
    except (ValueError, TypeError):
        pass
    s = str(s).strip().upper()
    s = re.sub(r'\s*-\s*', '-', s)
    return re.sub(r'\s+', ' ', s).strip()

def is_data_row(nombre):
    if not nombre:
        return False
    s = str(nombre).strip().upper()
    if len(s) < 3:
        return False
    prefixes = ('TOTAL', 'SUBTOTAL', 'SUMA', 'MONTO', 'N°', 'NOMBRE',
                'TITULAR', 'APELLIDOS', 'SOCIOS', 'INQUILINOS', 'COBRAR',
                'CONSUMO', 'LECTURA', 'PARTE CON', 'PARTE SIN',
                'LUZ DEL', 'LUZ DE ', 'RECIBO', 'PERIODO')
    for p in prefixes:
        if s.startswith(p):
            return False
    return True

def kwh_calc(amp, h=12, d=30):
    return (float(amp) * 220.0 * h * d) / 1000.0

def detect_horas_dias(kwh_excel, amp):
    """
    Detecta configuración de horas/dias comparando KWH excel con distintas combinaciones.
    Hallazgos del checkpoint:
      ratio ≈ 1.333 → dias_uso = 40 (período de 40 días)
      ratio ≈ 1.167 → horas_uso = 14 (14h de uso diario)
    """
    if not kwh_excel or amp <= 0:
        return 12, 30
    for dias in [40, 35, 30, 28]:
        for horas in [16, 14, 12]:
            expected = kwh_calc(amp, horas, dias)
            if expected > 0 and abs(kwh_excel / expected - 1.0) < 0.015:
                return horas, dias
    return 12, 30

def esc(s):
    return "'" + str(s).replace("'", "''") + "'"

def fmt_num(n, decimals=2):
    return f"{n:.{decimals}f}"

# ── 1. Luz Hoja1 — Amperaje (fuente canónica) ─────────────────────────────────
wb_luz = openpyxl.load_workbook(
    BASE / "Luz setiembre.xlsx", data_only=True, read_only=True
)

# Hoja1: R1=título, R2=header, data desde R3 (idx 2)
# cols: 0=N°, 1=Titular, 2=Tipo, 3=Puesto, 4=Amperaje, 5=KWH, 6=Precio, 7=SubTotal
amp_rows  = []   # (puesto, amp, horas, dias, precio, monto, tipo)
puestos_seen_amp = set()

ws = wb_luz['Hoja1']
for idx, row in enumerate(ws.iter_rows(values_only=True)):
    if idx < 2:
        continue
    if len(row) < 8:
        continue
    nombre, tipo = row[1], row[2]
    if not is_data_row(nombre):
        continue
    puesto = norm_puesto(row[3])
    if not puesto:
        continue
    try:
        amp   = float(row[4])
        kwh_e = float(row[5]) if row[5] is not None else None
        precio = float(row[6]) if row[6] is not None else 0.75
        subtotal = float(row[7])
    except (TypeError, ValueError):
        continue
    if amp <= 0 or subtotal <= 0:
        continue
    if puesto in puestos_seen_amp:
        continue
    horas, dias = detect_horas_dias(kwh_e, amp)
    tipo_u = str(tipo).strip().upper() if tipo else 'SOCIO'
    amp_rows.append((puesto, amp, horas, dias, precio, round(subtotal, 2), tipo_u))
    puestos_seen_amp.add(puesto)

non_std_amp = [(p, h, d) for p, _, h, d, _, _, _ in amp_rows if (h, d) != (12, 30)]

# ── 2. Luz CON MEDIDORES SOCIOS — registro_artefactos ────────────────────────
# R1=título, R2=precio global, R3=header, data desde R4 (idx 3)
# cols: 0=N°, 1=Nombre, 2=Puesto, 3=Lect_Ago, 4=Lect_Set, 5=KWH, 6=LuzTotal
med_socio_rows = []
puestos_seen_med_s = set()

ws = wb_luz[' SETIEMBRE CON MEDIDORES SOCIOS']
for idx, row in enumerate(ws.iter_rows(values_only=True)):
    if idx < 3:
        continue
    if len(row) < 7:
        continue
    nombre = row[1]
    if not is_data_row(nombre):
        continue
    puesto = norm_puesto(row[2])
    if not puesto or puesto in puestos_seen_med_s:
        continue
    try:
        kwh   = float(row[5]) if row[5] else 0.0
        monto = float(row[6])
    except (TypeError, ValueError):
        continue
    if monto <= 0:
        continue
    med_socio_rows.append((puesto, round(kwh, 3), round(monto, 2)))
    puestos_seen_med_s.add(puesto)

# ── 3. Luz INQUILINOS SETIEMBRE — registro_artefactos ────────────────────────
# R1=precio global (col 7 = 0.77), R2=header, data desde R3 (idx 2)
# cols: 0=N°, 1=Nombre, 2=Puesto, 3=Medidor, 4=Lect_Ago, 5=Lect_Set, 6=KWH, 7=SubTotal
med_inq_rows = []
puestos_seen_med_i = set()

ws = wb_luz['INQUILINOS SETIEMBRE']
for idx, row in enumerate(ws.iter_rows(values_only=True)):
    if idx < 2:
        continue
    if len(row) < 8:
        continue
    nombre = row[1]
    if not is_data_row(nombre):
        continue
    puesto = norm_puesto(row[2])
    if not puesto or puesto in puestos_seen_med_i:
        continue
    try:
        kwh   = float(row[6]) if row[6] else 0.0
        monto = float(row[7])
    except (TypeError, ValueError):
        continue
    if monto <= 0:
        continue
    med_inq_rows.append((puesto, round(kwh, 3), round(monto, 2)))
    puestos_seen_med_i.add(puesto)

wb_luz.close()

# Eliminar solapamientos: si un puesto está en medidores, no lo ponemos en amperaje
puestos_con_medidor = puestos_seen_med_s | puestos_seen_med_i
amp_rows = [(p, a, h, d, pr, m, t) for p, a, h, d, pr, m, t in amp_rows
            if p not in puestos_con_medidor]

# ── 4. Agua CON MEDIDORES — montos directos ───────────────────────────────────
# R1=header, data desde R2 (idx 1)
# cols: 0=N°, 1=Nombre, 2=Puesto, 3=N°Medidor, 4=Lect_Set, 5=Lect_Oct, 6=Diff, 7=Costo/m3, 8=Total
wb_agua = openpyxl.load_workbook(
    BASE / "AGUA octubre (1).xlsx", data_only=True, read_only=True
)

agua_medidores = []   # (puesto, monto)
puestos_agua_med = set()

ws = wb_agua['AGUA OCTUBRE CON MEDIDORES']
for idx, row in enumerate(ws.iter_rows(values_only=True)):
    if idx < 1:
        continue
    if len(row) < 9:
        continue
    nombre = row[1]
    if not is_data_row(nombre):
        continue
    puesto = norm_puesto(row[2])
    if not puesto or puesto in puestos_agua_med:
        continue
    try:
        monto = float(row[8])
    except (TypeError, ValueError):
        continue
    if monto <= 0:
        continue
    agua_medidores.append((puesto, round(monto, 2)))
    puestos_agua_med.add(puesto)

# ── 5. Agua Hoja1 — monto fijo, cruce por nombre ─────────────────────────────
# R5=header(idx4), R6=blank(idx5), data desde R7 (idx 6)
# cols: 0=N°, 1=Nombre, 2=Consumo base, 3=Ajuste, 4=Total, 5=Redondeo
agua_fija = []   # (nombre_display, monto)
nombres_agua_fija = set()

ws = wb_agua['Hoja1']
for idx, row in enumerate(ws.iter_rows(values_only=True)):
    if idx < 6:
        continue
    if len(row) < 5:
        continue
    nombre = row[1]
    if not is_data_row(nombre):
        continue
    nombre_str = str(nombre).strip()
    if nombre_str in nombres_agua_fija:
        continue
    try:
        # Preferir REDONDEO (col 5) si existe, sino TOTAL (col 4)
        monto_raw = row[5] if (len(row) > 5 and row[5]) else row[4]
        monto = float(monto_raw)
    except (TypeError, ValueError):
        continue
    if monto <= 0:
        continue
    agua_fija.append((nombre_str, round(monto, 2)))
    nombres_agua_fija.add(nombre_str)

wb_agua.close()

# ── Stats ─────────────────────────────────────────────────────────────────────
total_luz_inserts  = len(amp_rows) + len(med_socio_rows) + len(med_inq_rows)
total_agua_inserts = len(agua_medidores) + len(agua_fija)
total_inserts      = total_luz_inserts + total_agua_inserts

print(f"── Resumen de datos extraídos ──────────────────────────────────────")
print(f"Luz amperaje (Hoja1, deduped):          {len(amp_rows):>4}")
print(f"  Puestos con config no estándar:        {len([r for r in amp_rows if (r[2],r[3])!=(12,30)]):>4}  "
      f"{[(p, f'{h}h/{d}d') for p,_,h,d,_,_,_ in amp_rows if (h,d)!=(12,30)]}")
print(f"Luz con medidor socios:                 {len(med_socio_rows):>4}")
print(f"Luz con medidor inquilinos:             {len(med_inq_rows):>4}")
print(f"Agua con medidor (con código puesto):   {len(agua_medidores):>4}")
print(f"Agua fija (cruce por nombre en SQL):    {len(agua_fija):>4}")
print(f"──────────────────────────────────────────────────────────────────")
print(f"Total inserts esperados en montos_por_cobrar: {total_inserts}")
print()

# ═════════════════════════════════════════════════════════════════════════════
# GENERACIÓN SQL
# ═════════════════════════════════════════════════════════════════════════════

lines = []
A = lines.append

A("-- =============================================================================")
A("-- Migración 00007 — Seed de deudas iniciales (Luz Setiembre 2025, Agua Octubre 2025)")
A("-- Cooperativa Primero de Mayo · SistemaCooperativa")
A("-- -----------------------------------------------------------------------------")
A("-- Fuentes:")
A("--   • Luz:  Luz setiembre.xlsx  (Hoja1=amperaje, CON_MEDIDORES, INQUILINOS)")
A("--   • Agua: AGUA octubre (1).xlsx  (CON_MEDIDORES, Hoja1=monto fijo)")
A("--")
A(f"-- Inserts esperados:")
A(f"--   • {len(amp_rows):>3} mediciones_luz   (método amperaje, Hoja1)")
A(f"--   • {len(med_socio_rows)+len(med_inq_rows):>3} registro_artefactos (método medidor)")
A(f"--   • {total_inserts:>3} montos_por_cobrar (luz set-2025 + agua oct-2025)")
A("--")
A("-- Parámetros aplicados:")
A("--   • precio_kwh socios:     0.75 S/kWh  (Hoja1, consistente con subtotales)")
A("--   • precio_kwh inquilinos: 0.77 S/kWh  (INQUILINOS SETIEMBRE, encabezado)")
A("--   • monto_alumbrado:       0   (no columna separada detectada en ninguna hoja)")
A(f"--   • puestos con config especial: {[(p,f'{h}h/{d}d') for p,_,h,d,_,_,_ in amp_rows if (h,d)!=(12,30)]}")
A("--")
A("-- Idempotencia:")
A("--   • ON CONFLICT DO NOTHING en mediciones_luz, registro_artefactos, montos_por_cobrar")
A("--     via índices únicos parciales (puesto_id, periodo_anio, periodo_mes) WHERE deleted_at IS NULL")
A("-- =============================================================================")
A("")

# ─────────────────────────────────────────────────────────────────────────────
# SECCIÓN 1: mediciones_luz (amperaje, Hoja1)
# ─────────────────────────────────────────────────────────────────────────────
if amp_rows:
    A("-- -----------------------------------------------------------------------------")
    A(f"-- 1a. mediciones_luz — Amperaje ({len(amp_rows)} puestos, período 2025-09)")
    A("-- -----------------------------------------------------------------------------")
    A("with source_amp(codigo_puesto, amperaje, horas_uso, dias_uso, precio_kwh) as (values")
    amp_vals = [f"    ({esc(p)}, {fmt_num(a,3)}, {h}, {d}, {fmt_num(pr,4)})"
                for p, a, h, d, pr, _, _ in amp_rows]
    A(",\n".join(amp_vals))
    A(")")
    A("insert into public.mediciones_luz")
    A("  (puesto_id, periodo_anio, periodo_mes, fecha_medicion,")
    A("   amperaje, horas_uso, dias_uso, precio_kwh, monto_alumbrado)")
    A("select p.id, 2025, 9, '2025-09-30',")
    A("       s.amperaje, s.horas_uso, s.dias_uso, s.precio_kwh, 0")
    A("from source_amp s")
    A("join public.puestos p on p.codigo_puesto = s.codigo_puesto and p.deleted_at is null")
    A("on conflict (puesto_id, periodo_anio, periodo_mes) where deleted_at is null do nothing;")
    A("")

    A("-- 1b. montos_por_cobrar desde mediciones_luz amperaje")
    A("with source_amp(codigo_puesto, monto_asignado) as (values")
    amp_m_vals = [f"    ({esc(p)}, {fmt_num(m,2)})"
                  for p, _, _, _, _, m, _ in amp_rows]
    A(",\n".join(amp_m_vals))
    A(")")
    A("insert into public.montos_por_cobrar")
    A("  (puesto_id, concepto_id, periodo_anio, periodo_mes,")
    A("   monto, estado, metodo_calculo, medicion_luz_id, fecha_generacion)")
    A("select ml.puesto_id, c.id, 2025, 9,")
    A("       s.monto_asignado, 'Pendiente', 'Amperaje', ml.id, '2025-09-30'")
    A("from source_amp s")
    A("join public.puestos p on p.codigo_puesto = s.codigo_puesto and p.deleted_at is null")
    A("join public.mediciones_luz ml")
    A("  on ml.puesto_id = p.id and ml.periodo_anio = 2025 and ml.periodo_mes = 9")
    A("  and ml.deleted_at is null")
    A("join public.conceptos c on c.nombre = 'Luz' and c.deleted_at is null")
    A("on conflict (puesto_id, concepto_id, periodo_anio, periodo_mes)")
    A("  where deleted_at is null do nothing;")
    A("")

# ─────────────────────────────────────────────────────────────────────────────
# SECCIÓN 2: registro_artefactos + montos — socios con medidor
# ─────────────────────────────────────────────────────────────────────────────
if med_socio_rows:
    A("-- -----------------------------------------------------------------------------")
    A(f"-- 2a. registro_artefactos — Socios con medidor ({len(med_socio_rows)} puestos, período 2025-09)")
    A("-- -----------------------------------------------------------------------------")
    A("with source_med_s(codigo_puesto, kwh_consumido, monto_asignado) as (values")
    vals = [f"    ({esc(p)}, {fmt_num(k,3)}, {fmt_num(m,2)})"
            for p, k, m in med_socio_rows]
    A(",\n".join(vals))
    A(")")
    A("insert into public.registro_artefactos")
    A("  (puesto_id, periodo_anio, periodo_mes, fecha_registro,")
    A("   monto_asignado, otros_equipos)")
    A("select p.id, 2025, 9, '2025-09-30',")
    A("       s.monto_asignado, concat('KWH_consumido=', s.kwh_consumido::text)")
    A("from source_med_s s")
    A("join public.puestos p on p.codigo_puesto = s.codigo_puesto and p.deleted_at is null")
    A("on conflict (puesto_id, periodo_anio, periodo_mes) where deleted_at is null do nothing;")
    A("")

    A("-- 2b. montos_por_cobrar desde registro_artefactos socios")
    A("with source_med_s(codigo_puesto, monto_asignado) as (values")
    vals_m = [f"    ({esc(p)}, {fmt_num(m,2)})" for p, _, m in med_socio_rows]
    A(",\n".join(vals_m))
    A(")")
    A("insert into public.montos_por_cobrar")
    A("  (puesto_id, concepto_id, periodo_anio, periodo_mes,")
    A("   monto, estado, metodo_calculo, registro_artefacto_id, fecha_generacion)")
    A("select ra.puesto_id, c.id, 2025, 9,")
    A("       s.monto_asignado, 'Pendiente', 'Artefactos', ra.id, '2025-09-30'")
    A("from source_med_s s")
    A("join public.puestos p on p.codigo_puesto = s.codigo_puesto and p.deleted_at is null")
    A("join public.registro_artefactos ra")
    A("  on ra.puesto_id = p.id and ra.periodo_anio = 2025 and ra.periodo_mes = 9")
    A("  and ra.deleted_at is null")
    A("join public.conceptos c on c.nombre = 'Luz' and c.deleted_at is null")
    A("on conflict (puesto_id, concepto_id, periodo_anio, periodo_mes)")
    A("  where deleted_at is null do nothing;")
    A("")

# ─────────────────────────────────────────────────────────────────────────────
# SECCIÓN 3: registro_artefactos + montos — inquilinos con medidor
# ─────────────────────────────────────────────────────────────────────────────
if med_inq_rows:
    A("-- -----------------------------------------------------------------------------")
    A(f"-- 3a. registro_artefactos — Inquilinos con medidor ({len(med_inq_rows)} puestos, período 2025-09)")
    A("-- -----------------------------------------------------------------------------")
    A("with source_med_i(codigo_puesto, kwh_consumido, monto_asignado) as (values")
    vals = [f"    ({esc(p)}, {fmt_num(k,3)}, {fmt_num(m,2)})"
            for p, k, m in med_inq_rows]
    A(",\n".join(vals))
    A(")")
    A("insert into public.registro_artefactos")
    A("  (puesto_id, periodo_anio, periodo_mes, fecha_registro,")
    A("   monto_asignado, otros_equipos)")
    A("select p.id, 2025, 9, '2025-09-30',")
    A("       s.monto_asignado, concat('KWH_consumido=', s.kwh_consumido::text)")
    A("from source_med_i s")
    A("join public.puestos p on p.codigo_puesto = s.codigo_puesto and p.deleted_at is null")
    A("on conflict (puesto_id, periodo_anio, periodo_mes) where deleted_at is null do nothing;")
    A("")

    A("-- 3b. montos_por_cobrar desde registro_artefactos inquilinos")
    A("with source_med_i(codigo_puesto, monto_asignado) as (values")
    vals_m = [f"    ({esc(p)}, {fmt_num(m,2)})" for p, _, m in med_inq_rows]
    A(",\n".join(vals_m))
    A(")")
    A("insert into public.montos_por_cobrar")
    A("  (puesto_id, concepto_id, periodo_anio, periodo_mes,")
    A("   monto, estado, metodo_calculo, registro_artefacto_id, fecha_generacion)")
    A("select ra.puesto_id, c.id, 2025, 9,")
    A("       s.monto_asignado, 'Pendiente', 'Artefactos', ra.id, '2025-09-30'")
    A("from source_med_i s")
    A("join public.puestos p on p.codigo_puesto = s.codigo_puesto and p.deleted_at is null")
    A("join public.registro_artefactos ra")
    A("  on ra.puesto_id = p.id and ra.periodo_anio = 2025 and ra.periodo_mes = 9")
    A("  and ra.deleted_at is null")
    A("join public.conceptos c on c.nombre = 'Luz' and c.deleted_at is null")
    A("on conflict (puesto_id, concepto_id, periodo_anio, periodo_mes)")
    A("  where deleted_at is null do nothing;")
    A("")

# ─────────────────────────────────────────────────────────────────────────────
# SECCIÓN 4: Agua con medidor — montos directos (código puesto conocido)
# ─────────────────────────────────────────────────────────────────────────────
if agua_medidores:
    A("-- -----------------------------------------------------------------------------")
    A(f"-- 4. montos_por_cobrar — Agua con medidor ({len(agua_medidores)} puestos, período 2025-10)")
    A("-- -----------------------------------------------------------------------------")
    A("with source_agua_med(codigo_puesto, monto_asignado) as (values")
    vals = [f"    ({esc(p)}, {fmt_num(m,2)})" for p, m in agua_medidores]
    A(",\n".join(vals))
    A(")")
    A("insert into public.montos_por_cobrar")
    A("  (puesto_id, concepto_id, periodo_anio, periodo_mes,")
    A("   monto, estado, metodo_calculo, fecha_generacion)")
    A("select p.id, c.id, 2025, 10,")
    A("       s.monto_asignado, 'Pendiente', 'Fijo', '2025-10-31'")
    A("from source_agua_med s")
    A("join public.puestos p on p.codigo_puesto = s.codigo_puesto and p.deleted_at is null")
    A("join public.conceptos c on c.nombre = 'Agua' and c.deleted_at is null")
    A("on conflict (puesto_id, concepto_id, periodo_anio, periodo_mes)")
    A("  where deleted_at is null do nothing;")
    A("")

# ─────────────────────────────────────────────────────────────────────────────
# SECCIÓN 5: Agua fija — cruce por nombre (socios + inquilinos vía SQL JOIN)
# El JOIN contra socios.apellidos + historial_titularidad resuelve el puesto_id
# directamente en la base de datos.
# ─────────────────────────────────────────────────────────────────────────────
if agua_fija:
    A("-- -----------------------------------------------------------------------------")
    A(f"-- 5a. montos agua fija vía socios → titularidad ({len(agua_fija)} nombres, período 2025-10)")
    A("-- El JOIN cruza nombre con socios.apellidos + titularidad vigente → puesto_id.")
    A("-- Si no hay match, la fila se omite silenciosamente (INNER JOIN).")
    A("-- ON CONFLICT evita duplicados con los puestos ya insertados en sección 4.")
    A("-- -----------------------------------------------------------------------------")
    A("with source_agua_fija(nombre_display, monto_asignado) as (values")
    vals = [f"    ({esc(n)}, {fmt_num(m,2)})" for n, m in agua_fija]
    A(",\n".join(vals))
    A(")")
    A("insert into public.montos_por_cobrar")
    A("  (puesto_id, concepto_id, periodo_anio, periodo_mes,")
    A("   monto, estado, metodo_calculo, fecha_generacion)")
    A("select ht.puesto_id, c.id, 2025, 10,")
    A("       s.monto_asignado, 'Pendiente', 'Fijo', '2025-10-31'")
    A("from source_agua_fija s")
    A("join public.socios soc")
    A("  on soc.apellidos = s.nombre_display and soc.deleted_at is null")
    A("join public.historial_titularidad ht")
    A("  on ht.socio_id = soc.id and ht.fecha_fin is null")
    A("join public.conceptos c on c.nombre = 'Agua' and c.deleted_at is null")
    A("on conflict (puesto_id, concepto_id, periodo_anio, periodo_mes)")
    A("  where deleted_at is null do nothing;")
    A("")

    A("-- 5b. misma fuente vía inquilinos → arriendo vigente → puesto_id")
    A("with source_agua_fija(nombre_display, monto_asignado) as (values")
    A(",\n".join(vals))
    A(")")
    A("insert into public.montos_por_cobrar")
    A("  (puesto_id, concepto_id, periodo_anio, periodo_mes,")
    A("   monto, estado, metodo_calculo, fecha_generacion)")
    A("select ha.puesto_id, c.id, 2025, 10,")
    A("       s.monto_asignado, 'Pendiente', 'Fijo', '2025-10-31'")
    A("from source_agua_fija s")
    A("join public.inquilinos inq")
    A("  on inq.apellidos = s.nombre_display and inq.deleted_at is null")
    A("join public.historial_arriendos ha")
    A("  on ha.inquilino_id = inq.id and ha.fecha_fin is null")
    A("join public.conceptos c on c.nombre = 'Agua' and c.deleted_at is null")
    A("on conflict (puesto_id, concepto_id, periodo_anio, periodo_mes)")
    A("  where deleted_at is null do nothing;")
    A("")

# ─────────────────────────────────────────────────────────────────────────────
# SECCIÓN 6: Verificación final (query de conteo)
# ─────────────────────────────────────────────────────────────────────────────
A("-- -----------------------------------------------------------------------------")
A("-- 6. Verificación post-seed (ejecutar en SQL Editor para validar)")
A("-- -----------------------------------------------------------------------------")
A("-- select 'mediciones_luz (set-2025)'  as tabla, count(*) from public.mediciones_luz")
A("--   where periodo_anio=2025 and periodo_mes=9  and deleted_at is null")
A("-- union all")
A("-- select 'registro_artefactos (set-2025)', count(*) from public.registro_artefactos")
A("--   where periodo_anio=2025 and periodo_mes=9  and deleted_at is null")
A("-- union all")
A("-- select 'montos luz (set-2025)',  count(*) from public.montos_por_cobrar m")
A("--   join public.conceptos c on c.id=m.concepto_id")
A("--   where m.periodo_anio=2025 and m.periodo_mes=9 and m.deleted_at is null and c.nombre='Luz'")
A("-- union all")
A("-- select 'montos agua (oct-2025)', count(*) from public.montos_por_cobrar m")
A("--   join public.conceptos c on c.id=m.concepto_id")
A("--   where m.periodo_anio=2025 and m.periodo_mes=10 and m.deleted_at is null and c.nombre='Agua';")
A(f"--")
A(f"-- Esperado:")
A(f"--   mediciones_luz set-2025  : {len(amp_rows)}")
A(f"--   registro_artefactos      : {len(med_socio_rows)+len(med_inq_rows)}")
A(f"--   montos luz set-2025      : {len(amp_rows)+len(med_socio_rows)+len(med_inq_rows)}")
A(f"--   montos agua oct-2025     : ~{len(agua_medidores)+len(agua_fija)} (los que matcheen por nombre)")

sql_path = Path(r"c:/Github/SistemaCooperativa/supabase/migrations/00007_seed_deudas_iniciales.sql")
sql_path.write_text("\n".join(lines), encoding='utf-8')
print(f"✅ SQL generado → {sql_path}")
print(f"   Tamaño: {sql_path.stat().st_size:,} bytes  |  Líneas: {len(lines):,}")
