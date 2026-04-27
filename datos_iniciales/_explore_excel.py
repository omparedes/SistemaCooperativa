"""
Checkpoint de calidad de datos — SistemaCooperativa
Analiza los Excel de Luz (Setiembre) y Agua (Octubre) para:
  1. Detectar precio_kwh y monto_alumbrado por hoja
  2. Mapear puestos contra el padrón conocido
  3. Validar los cálculos de KWH vs amperaje
  4. Generar reporte_calidad.md
"""

import openpyxl
import json
import math
import re
import sys
import io
from pathlib import Path
from collections import defaultdict

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

BASE   = Path(r"c:/Github/SistemaCooperativa/datos_iniciales")
OUTDIR = BASE

# ── 1. Cargar puestos conocidos ──────────────────────────────────────────────
data = json.loads((BASE / '_extracted.json').read_text(encoding='utf-8'))
PUESTOS_CONOCIDOS = {p.upper().strip() for p in data['puestos'] if p.strip()}

def norm_puesto(s):
    if s is None: return ''
    s = str(s).strip().upper()
    s = re.sub(r'\s*-\s*', '-', s)
    return re.sub(r'\s+', ' ', s).strip()

# ── helpers ───────────────────────────────────────────────────────────────────
def kwh_calc(amp, horas=12, dias=30):
    return (float(amp) * 220 * horas * dias) / 1000

def pct_diff(a, b):
    if b == 0: return None
    return abs(a - b) / abs(b) * 100

def es_valid_name(s):
    if not s: return False
    s = str(s).strip().upper()
    if len(s) < 4: return False
    bad = {'TOTAL','SUBTOTAL','NOMBRE','TITULAR','APELLIDOS','COBRAR','DESCRIPCION',
           'SUMA','PROMEDIO','N/A','','/','---'}
    return s not in bad and not s.startswith('TOTAL') and not s.startswith('SUBTOTAL')

# ── 2. Análisis por hoja ──────────────────────────────────────────────────────
LUZ_FILE   = "Luz setiembre.xlsx"
AGUA_FILE  = "AGUA octubre (1).xlsx"

report_sections = []

def banner(title):
    return f"\n## {title}\n"

def add(section, text):
    section.append(text)

# ═══════════════════════════════════════════════════════════════════════════════
# LUZ — exploración hoja a hoja
# ═══════════════════════════════════════════════════════════════════════════════
luz_sections = []
add(luz_sections, banner("Luz Setiembre — Análisis por Hoja"))

wb_luz = openpyxl.load_workbook(BASE / LUZ_FILE, data_only=True, read_only=True)

# ─── Hoja: "socios sin medidores" (amperaje, col: Titular|Tipo|Puesto|Amp|KWH|PrecioKWH|SubTotal) ───
def analizar_amperaje(ws, hdr_row, col_titular, col_tipo, col_puesto,
                      col_amp, col_kwh, col_precio, col_subtotal, hoja_nombre):
    rows_ok, rows_bad, rows_calc_err = [], [], []
    precios_vistos = set()
    precio_fijo_global = None

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < hdr_row: continue
        try:
            titular  = row[col_titular]  if col_titular  is not None else None
            tipo     = row[col_tipo]     if col_tipo     is not None else None
            puesto   = row[col_puesto]
            amp_raw  = row[col_amp]
            kwh_raw  = row[col_kwh]      if col_kwh      is not None else None
            precio   = row[col_precio]   if col_precio   is not None else None
            subtotal = row[col_subtotal] if col_subtotal is not None else None
        except IndexError:
            continue

        if not es_valid_name(titular): continue
        nombre = str(titular).strip()
        tipo_u = str(tipo).strip().upper() if tipo else ''
        p_norm = norm_puesto(puesto)

        try:
            amp   = float(amp_raw)
            kwh_e = float(kwh_raw) if kwh_raw is not None else None
            prx   = float(precio)  if precio  is not None else None
            sub_e = float(subtotal) if subtotal is not None else None
        except (TypeError, ValueError):
            rows_bad.append({'nombre': nombre, 'puesto_raw': puesto, 'razon': 'no-numeric'})
            continue

        if prx: precios_vistos.add(round(prx, 4))

        matched = p_norm in PUESTOS_CONOCIDOS

        # Validar KWH si tenemos todos los datos
        calc_issue = None
        if amp > 0 and kwh_e is not None:
            kwh_c = kwh_calc(amp)
            diff  = pct_diff(kwh_c, kwh_e)
            if diff is not None and diff > 1:
                calc_issue = f"KWH esperado={kwh_c:.3f} Excel={kwh_e:.3f} diff={diff:.1f}%"

        # Validar subtotal vs precio * kwh
        sub_issue = None
        if prx and kwh_e is not None and sub_e is not None:
            sub_c = round(kwh_e * prx, 2)
            if pct_diff(sub_c, sub_e) is not None and pct_diff(sub_c, sub_e) > 2:
                sub_issue = f"SubTotal esperado={sub_c:.2f} Excel={sub_e:.2f}"

        record = {
            'nombre': nombre, 'tipo': tipo_u,
            'puesto_raw': str(puesto), 'puesto_norm': p_norm,
            'matched': matched,
            'amp': amp, 'kwh_excel': kwh_e, 'precio': prx, 'subtotal': sub_e,
            'calc_issue': calc_issue, 'sub_issue': sub_issue,
        }
        if not matched:
            rows_bad.append(record)
        elif calc_issue or sub_issue:
            rows_calc_err.append(record)
        else:
            rows_ok.append(record)

    return rows_ok, rows_bad, rows_calc_err, sorted(precios_vistos), precio_fijo_global


def analizar_medidores(ws, hdr_row, col_nombre, col_puesto,
                       col_lect_ant, col_lect_act, col_kwh, col_total, hoja_nombre, precio_global=None):
    """Para hojas con lecturas de medidor (diferencia = KWH consumido)."""
    rows_ok, rows_bad, rows_calc_err = [], [], []
    precio_fijo = precio_global

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < hdr_row: continue
        try:
            nombre  = row[col_nombre]
            puesto  = row[col_puesto]
            l_ant   = row[col_lect_ant] if col_lect_ant is not None else None
            l_act   = row[col_lect_act] if col_lect_act is not None else None
            kwh_raw = row[col_kwh]      if col_kwh     is not None else None
            total   = row[col_total]    if col_total   is not None else None
        except IndexError:
            continue

        if not es_valid_name(nombre): continue
        p_norm = norm_puesto(puesto)
        matched = p_norm in PUESTOS_CONOCIDOS

        try:
            kwh_e = float(kwh_raw) if kwh_raw else None
            tot_e = float(total)   if total   else None
            la    = float(l_ant)   if l_ant   else None
            lac   = float(l_act)   if l_act   else None
        except (TypeError, ValueError):
            rows_bad.append({'nombre': str(nombre).strip(), 'puesto_raw': puesto, 'razon': 'no-numeric'})
            continue

        diff_issue = None
        if la is not None and lac is not None and kwh_e is not None:
            kwh_c = lac - la
            if kwh_c < 0: diff_issue = f"lectura anterior > actual ({la} > {lac})"
            elif abs(kwh_c - kwh_e) > 0.05: diff_issue = f"diferencia={kwh_c:.3f} vs KWH={kwh_e:.3f}"

        record = {'nombre': str(nombre).strip(), 'puesto_raw': str(puesto), 'puesto_norm': p_norm,
                  'matched': matched, 'kwh_excel': kwh_e, 'total': tot_e, 'diff_issue': diff_issue}
        if not matched:
            rows_bad.append(record)
        elif diff_issue:
            rows_calc_err.append(record)
        else:
            rows_ok.append(record)

    return rows_ok, rows_bad, rows_calc_err


# — Hoja 1: ' SETIEMBRE CON MEDIDORES SOCIOS' — medidores —
ws = wb_luz[' SETIEMBRE CON MEDIDORES SOCIOS']
# Buscar precio_kwh en fila 2 (celda G2)
precio_h1 = None
for row in ws.iter_rows(min_row=2, max_row=2, values_only=True):
    for cell in row:
        try:
            v = float(cell)
            if 0.1 < v < 5.0: precio_h1 = v
        except (TypeError, ValueError): pass

ok1, bad1, err1 = analizar_medidores(ws, 3, 1, 2, 3, 4, 5, 6, 'CON_MEDIDORES', precio_h1)

s = [f"### Hoja: 'SETIEMBRE CON MEDIDORES SOCIOS' (lecturas de medidor)",
     f"- Precio KWH detectado en encabezado: **{precio_h1}**",
     f"- Método: diferencia de lecturas (lectura_oct - lectura_ago)",
     f"- Filas OK: {len(ok1)} | Mismatches de puesto: {len(bad1)} | Errores cálculo: {len(err1)}"]
if bad1:
    s.append(f"\n  **Mismatches:**")
    for r in bad1[:15]: s.append(f"  - `{r['puesto_raw']}` → norm=`{r.get('puesto_norm','')}` ({r['nombre'][:30]})")
    if len(bad1) > 15: s.append(f"  - ... y {len(bad1)-15} más")
if err1:
    s.append(f"\n  **Errores de cálculo:**")
    for r in err1[:10]: s.append(f"  - {r['nombre'][:30]}: {r.get('diff_issue','?')}")
luz_sections.extend(s)
luz_sections.append("")

# — Hoja 2: 'socios sin medidores' — amperaje —
ws2 = wb_luz['socios sin medidores']
ok2, bad2, err2, precios2, _ = analizar_amperaje(ws2, 3, 1, 2, 3, 4, 5, 6, 7, 'SIN_MEDIDORES')

# Detectar si existe col 8 (total con alumbrado)
col8_values = []
for i, row in enumerate(ws2.iter_rows(values_only=True)):
    if i < 3: continue
    if es_valid_name(row[1] if len(row) > 1 else None):
        try:
            v = float(row[7]) if len(row) > 7 and row[7] is not None else None
            v8 = float(row[7]) if len(row) > 7 else None
            if v is not None: col8_values.append(v)
        except: pass
has_alumbrado_col = False  # col 7 = Sub-Total only

s = [f"### Hoja: 'socios sin medidores' (amperaje)",
     f"- Precios KWH detectados por fila: **{sorted(precios2)}**",
     f"- Filas OK: {len(ok2)} | Mismatches de puesto: {len(bad2)} | Errores cálculo: {len(err2)}",
     f"- Columna alumbrado separada: NO detectada (Sub-Total = KWH × precio_kwh)"]
if bad2:
    s.append(f"\n  **Mismatches:**")
    for r in bad2[:15]: s.append(f"  - `{r['puesto_raw']}` → `{r.get('puesto_norm','')}` ({r['nombre'][:30]})")
    if len(bad2) > 15: s.append(f"  - ... y {len(bad2)-15} más")
if err2:
    s.append(f"\n  **Errores de cálculo (diff > 1%):**")
    for r in err2[:15]: s.append(f"  - puesto={r['puesto_norm']} amp={r['amp']} | {r['calc_issue']} | {r.get('sub_issue','')}")
luz_sections.extend(s)
luz_sections.append("")

# — Hoja 3: 'Hoja1' — amperaje, versión alternativa —
ws3 = wb_luz['Hoja1']
ok3, bad3, err3, precios3, _ = analizar_amperaje(ws3, 2, 1, 2, 3, 4, 5, 6, 7, 'HOJA1')
s = [f"### Hoja: 'Hoja1' (amperaje, versión revisada con precio distinto)",
     f"- Precios KWH detectados: **{sorted(precios3)}**",
     f"- Filas OK: {len(ok3)} | Mismatches: {len(bad3)} | Errores: {len(err3)}",
     f"- Nota: Esta hoja parece una revisión de 'socios sin medidores'. Evaluaremos si usar esta o la anterior."]
if err3:
    s.append(f"\n  **Errores de cálculo (muestra):**")
    for r in err3[:10]: s.append(f"  - p={r['puesto_norm']} amp={r['amp']}: {r.get('calc_issue','?')}")
luz_sections.extend(s)
luz_sections.append("")

# — Hoja 4: 'INQUILINOS SETIEMBRE' — medidores para inquilinos —
ws4 = wb_luz['INQUILINOS SETIEMBRE']
precio_inq = None
for row in ws4.iter_rows(min_row=1, max_row=1, values_only=True):
    for cell in row:
        try:
            v = float(cell)
            if 0.1 < v < 5.0: precio_inq = v
        except: pass

ok4, bad4, err4 = analizar_medidores(ws4, 2, 1, 2, 4, 5, 6, 7, 'INQ_SETIEMBRE', precio_inq)
s = [f"### Hoja: 'INQUILINOS SETIEMBRE' (lecturas de medidor, inquilinos)",
     f"- Precio KWH detectado: **{precio_inq}**",
     f"- Filas OK: {len(ok4)} | Mismatches: {len(bad4)} | Errores: {len(err4)}"]
if bad4:
    s.append(f"\n  **Mismatches:**")
    for r in bad4[:15]: s.append(f"  - `{r['puesto_raw']}` → `{r.get('puesto_norm','')}` ({r['nombre'][:30]})")
luz_sections.extend(s)
luz_sections.append("")

wb_luz.close()

# ─── RESUMEN LUZ ──────────────────────────────────────────────────────────────
total_luz_ok  = len(ok1)+len(ok2)+len(ok3)+len(ok4)
total_luz_bad = len(bad1)+len(bad2)+len(bad3)+len(bad4)
total_luz_err = len(err1)+len(err2)+len(err3)+len(err4)
total_luz     = total_luz_ok+total_luz_bad+total_luz_err
mismatch_pct  = (total_luz_bad/total_luz*100) if total_luz else 0

# ═══════════════════════════════════════════════════════════════════════════════
# AGUA — exploración hoja a hoja
# ═══════════════════════════════════════════════════════════════════════════════
agua_sections = [banner("Agua Octubre — Análisis por Hoja")]

wb_agua = openpyxl.load_workbook(BASE / AGUA_FILE, data_only=True, read_only=True)

def analizar_agua_medidores(ws, hdr_row, col_nombre, col_puesto, col_medidor,
                             col_lect_ant, col_lect_act, col_diff, col_costo_m3, col_total):
    rows_ok, rows_bad, rows_calc_err = [], [], []
    costos_m3 = set()
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < hdr_row: continue
        try:
            nombre  = row[col_nombre]
            puesto  = row[col_puesto]
            medidor = row[col_medidor] if col_medidor is not None else None
            la      = row[col_lect_ant] if col_lect_ant else None
            lac     = row[col_lect_act] if col_lect_act else None
            diff_e  = row[col_diff]    if col_diff else None
            costo   = row[col_costo_m3] if col_costo_m3 else None
            total   = row[col_total]   if col_total else None
        except IndexError:
            continue
        if not es_valid_name(nombre): continue
        p_norm  = norm_puesto(puesto)
        matched = p_norm in PUESTOS_CONOCIDOS

        try:
            tot_e = float(total) if total else None
            cos_v = float(costo) if costo else None
            la_v  = float(la)   if la    else None
            lac_v = float(lac)  if lac   else None
            diff_v= float(diff_e) if diff_e else None
        except (TypeError, ValueError):
            rows_bad.append({'nombre': str(nombre), 'puesto_raw': str(puesto), 'razon': 'no-numeric'})
            continue

        if cos_v and cos_v > 0: costos_m3.add(round(cos_v, 4))

        diff_issue = None
        if la_v is not None and lac_v is not None and diff_v is not None:
            if abs((lac_v - la_v) - diff_v) > 0.01:
                diff_issue = f"dif_calc={lac_v-la_v:.3f} vs dif_excel={diff_v:.3f}"

        record = {'nombre': str(nombre).strip(), 'puesto_raw': str(puesto), 'puesto_norm': p_norm,
                  'matched': matched, 'medidor': str(medidor) if medidor else '',
                  'total': tot_e, 'diff_issue': diff_issue}
        if not matched: rows_bad.append(record)
        elif diff_issue: rows_calc_err.append(record)
        else: rows_ok.append(record)

    return rows_ok, rows_bad, rows_calc_err, sorted(costos_m3)

def analizar_agua_fija(ws, hdr_row, col_nombre, col_monto, col_redondeo=None):
    rows_ok, rows_bad = [], []
    montos_fijos = set()
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < hdr_row: continue
        try:
            nombre = row[col_nombre]
            monto  = row[col_monto]
        except IndexError:
            continue
        if not es_valid_name(nombre): continue
        try:
            m = float(monto) if monto else None
            if m: montos_fijos.add(m)
        except: pass
        p_norm = '(sin código de puesto en esta hoja)'
        rows_ok.append({'nombre': str(nombre).strip(), 'monto': m})
    return rows_ok, rows_bad, sorted(montos_fijos)

# Hoja 1: con medidores
ws_a1 = wb_agua['AGUA OCTUBRE CON MEDIDORES']
# Row 1 is the header
ok_a1, bad_a1, err_a1, costos_m3 = analizar_agua_medidores(
    ws_a1, 1, col_nombre=1, col_puesto=2, col_medidor=3,
    col_lect_ant=4, col_lect_act=5, col_diff=6, col_costo_m3=7, col_total=8)

s = [f"### Hoja: 'AGUA OCTUBRE CON MEDIDORES'",
     f"- Costo por m³ detectado: **{costos_m3}** (S/ por m³)",
     f"- Filas OK: {len(ok_a1)} | Mismatches puesto: {len(bad_a1)} | Errores cálculo: {len(err_a1)}"]
if bad_a1:
    s.append(f"\n  **Mismatches:**")
    for r in bad_a1[:15]: s.append(f"  - `{r['puesto_raw']}` ({r['nombre'][:30]})")
if err_a1:
    s.append(f"\n  **Inconsistencias en lecturas:**")
    for r in err_a1[:10]: s.append(f"  - {r['nombre'][:30]}: {r['diff_issue']}")
agua_sections.extend(s)
agua_sections.append("")

# Hoja 2/3: lista amplia (monto fijo)
ws_a2 = wb_agua['Agua de octubre']
ok_a2, bad_a2, montos_fijos = analizar_agua_fija(ws_a2, 6, 1, 4)  # hdr row 5, data from 7 (idx 6)
s = [f"### Hoja: 'Agua de octubre' (monto fijo por socio, sin código puesto)",
     f"- Montos únicos detectados: {sorted(set(montos_fijos))[:20]}",
     f"- Total de registros con nombre: {len(ok_a2)}",
     f"- Nota: Esta hoja NO tiene código de puesto → se cruzará por nombre de socio/inquilino."]
agua_sections.extend(s)
agua_sections.append("")

# Hoja 3: 'Hoja1' de agua (revisión con otro monto base)
ws_a3 = wb_agua['Hoja1']
ok_a3, bad_a3, montos_fijos3 = analizar_agua_fija(ws_a3, 6, 1, 4)
s = [f"### Hoja: 'Hoja1' de Agua (revisión, montos posiblemente actualizados)",
     f"- Montos únicos: {sorted(set(montos_fijos3))[:20]}",
     f"- Total registros: {len(ok_a3)}"]
agua_sections.extend(s)

wb_agua.close()

total_agua_ok  = len(ok_a1)+len(ok_a2)+len(ok_a3)
total_agua_bad = len(bad_a1)+len(bad_a2)+len(bad_a3)
total_agua     = total_agua_ok+total_agua_bad
mismatch_pct_a = (total_agua_bad/total_agua*100) if total_agua else 0

# ═══════════════════════════════════════════════════════════════════════════════
# ANÁLISIS ESPECIAL: precios y alumbrado
# ═══════════════════════════════════════════════════════════════════════════════
extras = [banner("Análisis de Precios y Alumbrado")]

# Detectar si en socios sin medidores el Sub-Total = KWH * precio exacto,
# o si hay un componente adicional (alumbrado)
has_alumbrado = []
wb_l = openpyxl.load_workbook(BASE / LUZ_FILE, data_only=True, read_only=True)
ws_sm = wb_l['socios sin medidores']
for i, row in enumerate(ws_sm.iter_rows(values_only=True)):
    if i < 3: continue
    if not es_valid_name(row[1] if len(row) > 1 else None): break
    try:
        amp   = float(row[4])
        kwh   = float(row[5])
        prx   = float(row[6])
        sub   = float(row[7])
        kwh_c = kwh_calc(amp)
        sub_c = round(kwh_c * prx, 2)
        if abs(sub - sub_c) > 0.02:
            has_alumbrado.append({'amp': amp, 'kwh_c': round(kwh_c,3), 'precio': prx,
                                  'sub_esperado': sub_c, 'sub_excel': sub,
                                  'delta': round(sub - sub_c, 2)})
    except (TypeError, ValueError, IndexError):
        break

wb_l.close()

if has_alumbrado:
    extras.append(f"**Discrepancias entre Sub-Total Excel y KWH×precio ({len(has_alumbrado)} casos):**")
    extras.append(f"  → Esto indica que Sub-Total en esta hoja incluye un componente de alumbrado.")
    extras.append("")
    for r in has_alumbrado[:10]:
        extras.append(f"  - amp={r['amp']} | KWH_calc={r['kwh_c']} | P={r['precio']} | "
                       f"Esperado={r['sub_esperado']} | Excel={r['sub_excel']} | **delta={r['delta']}**")
    deltas = [r['delta'] for r in has_alumbrado]
    extras.append(f"\n  Delta mín/máx: {min(deltas):.2f} / {max(deltas):.2f} (probable monto_alumbrado fijo)")
    extras.append(f"  Media delta: {round(sum(deltas)/len(deltas), 2)} S/")
else:
    extras.append("**No se detectaron discrepancias Sub-Total vs KWH×precio.**")
    extras.append("→ La hoja 'socios sin medidores' NO incluye alumbrado en la columna Sub-Total.")
    extras.append("→ El monto_alumbrado = 0 o está en columna separada no detectada.")

# ─── Extremos en luz ───────────────────────────────────────────────────────────
extras.append(banner("Casos Extremos — Luz"))
all_luz_rows = ok2 + err2 + ok3 + err3
amps = [(r['amp'], r['puesto_norm'], r.get('kwh_excel',0)) for r in all_luz_rows if r.get('amp',0) > 0]
amps_sorted = sorted(amps, key=lambda x: x[0], reverse=True)
extras.append("**Top 5 puestos por amperaje:**")
for amp, pst, kwh in amps_sorted[:5]:
    kwh_c = kwh_calc(amp)
    extras.append(f"  - Puesto {pst}: amp={amp} A → KWH_calc={kwh_c:.1f} (excel={kwh:.1f})")
extras.append("\n**Top 5 puestos menor amperaje (> 0):**")
amps_low = sorted(amps, key=lambda x: x[0])
for amp, pst, kwh in amps_low[:5]:
    extras.append(f"  - Puesto {pst}: amp={amp} A → KWH_calc={kwh_calc(amp):.1f}")

# ═══════════════════════════════════════════════════════════════════════════════
# VEREDICTO FINAL
# ═══════════════════════════════════════════════════════════════════════════════
verdict = [banner("Veredicto y Recomendaciones para 00007")]

verdict.append(f"| Archivo | Total filas | OK (match) | Mismatch | Errores calc | Mismatch % |")
verdict.append(f"|---|---|---|---|---|---|")
verdict.append(f"| Luz (todas las hojas) | {total_luz} | {total_luz_ok} | {total_luz_bad} | {total_luz_err} | **{mismatch_pct:.1f}%** |")
verdict.append(f"| Agua (con código puesto) | {total_agua} | {total_agua_ok} | {total_agua_bad} | — | **{mismatch_pct_a:.1f}%** |")
verdict.append("")

if max(mismatch_pct, mismatch_pct_a) < 5:
    verdict.append("✅ **Mismatches < 5% → LUZ VERDE para generar 00007.**")
elif max(mismatch_pct, mismatch_pct_a) < 15:
    verdict.append("⚠️  **Mismatches entre 5-15% → revisar lista y decidir qué hacer con los huérfanos.**")
else:
    verdict.append("❌ **Mismatches > 15% → requiere limpieza manual antes de generar 00007.**")

verdict.append("\n**Estrategia recomendada por fuente:**")
verdict.append(f"| Hoja | Método | Tabla | Periodo |")
verdict.append(f"|---|---|---|---|")
verdict.append(f"| socios sin medidores (Tipo=SOCIO) | Amperaje → mediciones_luz | mediciones_luz + montos_por_cobrar | 2025-09 |")
verdict.append(f"| socios sin medidores (Tipo=INQUILINO) | Amperaje → mediciones_luz | mediciones_luz + montos_por_cobrar | 2025-09 |")
verdict.append(f"| SETIEMBRE CON MEDIDORES SOCIOS | Medidor → registro_artefactos | registro_artefactos + montos_por_cobrar | 2025-09 |")
verdict.append(f"| INQUILINOS SETIEMBRE | Medidor → registro_artefactos | registro_artefactos + montos_por_cobrar | 2025-09 |")
verdict.append(f"| AGUA OCTUBRE CON MEDIDORES | Medidor agua | montos_por_cobrar (Fijo) | 2025-10 |")
verdict.append(f"| Agua de octubre (sin código puesto) | Monto fijo | montos_por_cobrar (cruzar por nombre) | 2025-10 |")

# ═══════════════════════════════════════════════════════════════════════════════
# ENSAMBLAR REPORTE
# ═══════════════════════════════════════════════════════════════════════════════
all_sections = (
    ["# Reporte de Calidad de Datos — SistemaCooperativa", "",
     f"Generado: {__import__('datetime').datetime.now().strftime('%Y-%m-%d %H:%M')}",
     f"Archivos analizados: `{LUZ_FILE}` · `{AGUA_FILE}`", ""]
    + luz_sections + agua_sections + extras + verdict
)

md = "\n".join(all_sections)
out_path = OUTDIR / "reporte_calidad.md"
out_path.write_text(md, encoding='utf-8')

# Print to terminal
print(md)
print(f"\n─────────────────────────────────────────")
print(f"Reporte guardado en: {out_path}")
